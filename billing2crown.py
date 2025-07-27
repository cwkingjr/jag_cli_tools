# Using JAG billing export spreadsheet, generate an intermediate reformatted spreadsheet
# in a format conducive to loading into Crown's template, then using a copy of Crown's template,
# load that data into a new Crown billing spreadsheet.

import openpyxl

# TODO: Use the below snippets to work out how to get this done.

workbook = openpyxl.load_workbook("your_excel_file.xlsx")

# Select the desired worksheet by name (e.g., 'Sheet1')
# Or, select the active sheet if you don't know the name or it's the only one
sheet = workbook["Sheet1"]

sheet["A1"] = "New Value for A1"
sheet["B5"] = 123

# Save and overwrite the original file
workbook.save("your_excel_file.xlsx")

# Or, save as a new file
# workbook.save('new_excel_file.xlsx')

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook("example.xlsx")

# Select the sheet
sheet = workbook["Sheet1"]

# Write to specific cells
sheet["A1"] = "Hello"
sheet.cell(row=1, column=2).value = "World"
sheet["C3"] = 12345

# Save the changes
workbook.save("example.xlsx")
